#!/usr/bin/env python3
"""
DPSVM simulation and real-data experiments.

This script implements the RR-corrected convolution-smoothed SVM described in the user's
DPSVM draft. It runs:
1) a Gaussian class-conditional simulation with n=3000 and B=500 replications,
2) a real-data study on the Spambase dataset (N=4601) using PCA-reduced
   orthogonal features aligned with the fixed-p theory.
"""

import os
# Prevent BLAS over-subscription inside multiprocessing workers.
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("VECLIB_MAXIMUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")

from dataclasses import dataclass
import json
import math
import time
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import numpy as np
import pandas as pd
from scipy.optimize import brentq, minimize
from scipy.stats import norm

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sklearn.model_selection import StratifiedShuffleSplit
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA

# ---------------------------
# Global configuration
# ---------------------------
EPS_GRID: List[float] = [0.05, 0.06, 0.07, 0.08, 0.09, 0.10, 0.50, 0.70, 1.00]
DELTA_APPROX: float = 1e-4
ALPHA: float = 0.05
Z975: float = float(norm.ppf(1 - ALPHA / 2))

SIM_N: int = 3000
SIM_B: int = 500
SIM_WORKERS: int = min(8, max(4, (os.cpu_count() or 4) // 3))
REAL_SPLITS: int = 50
REAL_WORKERS: int = min(6, max(4, (os.cpu_count() or 4) // 4))

# ---------------------------
# Simulation Parameters
# ---------------------------
H_SCALE: float = 0.30
LAM_SCALE: float = 0.05
LOCAL_BOX_RADIUS: float = 0.80

# ---------------------------
# Real Data Parameters (Tuned for Spambase N~4600)
# ---------------------------
REAL_LOCAL_BOX_RADIUS: float = 4.0
REAL_H_SCALE: float = 0.80
REAL_LAM_SCALE: float = 0.10
REAL_PCA_COMPONENTS: int = 5

# Simulation scenario chosen after pilot tuning.
SIM_MU = np.array([1.2, 0.4, 1.0], dtype=float)
SIM_SIGMA = np.diag(np.array([0.8**2, 1.0**2, 0.7**2], dtype=float))

OUTPUT_DIR = Path("C:\\Users\\86182\\Desktop\\新建文件夹 (4)")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SQRT_2PI = math.sqrt(2.0 * math.pi)


# ---------------------------
# Core numerical utilities
# ---------------------------
def phi(z: np.ndarray) -> np.ndarray:
    return np.exp(-0.5 * z * z) / SQRT_2PI

def Phi(z: np.ndarray) -> np.ndarray:
    return norm.cdf(z)

def smoothed_hinge(u: np.ndarray, h: float) -> np.ndarray:
    z = u / h
    return u * Phi(z) + h * phi(z)

@dataclass
class FitResult:
    theta: np.ndarray
    loss: float
    hess: np.ndarray
    score_i: np.ndarray

def svm_components(theta: np.ndarray, X: np.ndarray, y: np.ndarray, h: float, lam: float) -> Tuple[float, np.ndarray, np.ndarray, np.ndarray]:
    z = X @ theta
    u = 1.0 - y * z
    cdf = Phi(u / h)
    pdf = phi(u / h)

    loss = float(smoothed_hinge(u, h).mean() + 0.5 * lam * np.dot(theta[1:], theta[1:]))
    grad = -(X * (y * cdf)[:, None]).mean(axis=0)
    grad[1:] += lam * theta[1:]

    hess = (X.T * (pdf / h)).dot(X) / X.shape[0]
    hess[1:, 1:] += lam * np.eye(X.shape[1] - 1)

    score_i = -(X * (y * cdf)[:, None])
    return loss, grad, hess, score_i

def dpsvm_components(theta: np.ndarray, X: np.ndarray, y_priv: np.ndarray, p_minus: float, p_plus: float, h: float, lam: float) -> Tuple[float, np.ndarray, np.ndarray, np.ndarray]:
    Delta = p_minus + p_plus - 1.0
    z = X @ theta
    u_minus = 1.0 - z
    u_plus = 1.0 + z

    cdf_minus = Phi(u_minus / h)
    cdf_plus = Phi(u_plus / h)
    pdf_minus = phi(u_minus / h)
    pdf_plus = phi(u_plus / h)

    loss_minus = smoothed_hinge(u_minus, h)
    loss_plus = smoothed_hinge(u_plus, h)

    pos = y_priv == 1
    neg = ~pos

    loss_i = np.empty_like(z)
    loss_i[pos] = (p_minus * loss_minus[pos] - (1.0 - p_plus) * loss_plus[pos]) / Delta
    loss_i[neg] = (p_plus * loss_plus[neg] - (1.0 - p_minus) * loss_minus[neg]) / Delta
    loss = float(loss_i.mean() + 0.5 * lam * np.dot(theta[1:], theta[1:]))

    weight = np.empty_like(z)
    weight[pos] = -(p_minus * cdf_minus[pos] + (1.0 - p_plus) * cdf_plus[pos]) / Delta
    weight[neg] = (p_plus * cdf_plus[neg] + (1.0 - p_minus) * cdf_minus[neg]) / Delta
    grad = (X * weight[:, None]).mean(axis=0)
    grad[1:] += lam * theta[1:]

    curv = np.empty_like(z)
    curv[pos] = (p_minus * pdf_minus[pos] - (1.0 - p_plus) * pdf_plus[pos]) / (h * Delta)
    curv[neg] = (p_plus * pdf_plus[neg] - (1.0 - p_minus) * pdf_minus[neg]) / (h * Delta)
    hess = (X.T * curv).dot(X) / X.shape[0]
    hess[1:, 1:] += lam * np.eye(X.shape[1] - 1)

    score_i = X * weight[:, None]
    return loss, grad, hess, score_i

def fit_lbfgs(
    components_func,
    theta0: np.ndarray,
    X: np.ndarray,
    y: np.ndarray,
    args: Tuple,
    bounds=None,
    maxiter: int = 300,
) -> FitResult:
    objective = lambda th: components_func(th, X, y, *args)[0]
    gradient = lambda th: components_func(th, X, y, *args)[1]
    res = minimize(
        objective,
        theta0,
        jac=gradient,
        method="L-BFGS-B",
        bounds=bounds,
        options={"maxiter": maxiter, "ftol": 1e-12, "gtol": 1e-8},
    )
    loss, _, hess, score_i = components_func(res.x, X, y, *args)
    return FitResult(theta=res.x, loss=float(loss), hess=hess, score_i=score_i)

def sandwich_covariance(hess: np.ndarray, score_i: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    n = score_i.shape[0]
    centered = score_i - score_i.mean(axis=0, keepdims=True)
    V_hat = centered.T @ centered / n
    H = 0.5 * (hess + hess.T)
    H_inv = np.linalg.pinv(H)
    Sigma_hat = H_inv @ V_hat @ H_inv
    return V_hat, Sigma_hat

def ci_coverage(theta_hat: np.ndarray, theta_true: np.ndarray, Sigma_hat: np.ndarray, n: int) -> float:
    se = np.sqrt(np.maximum(np.diag(Sigma_hat), 0.0) / n)
    covered = np.abs(theta_hat - theta_true) <= Z975 * se
    return float(np.mean(covered))

def privatize_labels(y_true: np.ndarray, p_minus: float, p_plus: float, rng: np.random.Generator) -> np.ndarray:
    y_priv = y_true.copy()
    neg = y_true == -1
    pos = ~neg
    y_priv[neg] = np.where(rng.random(neg.sum()) < p_minus, -1, 1)
    y_priv[pos] = np.where(rng.random(pos.sum()) < p_plus, 1, -1)
    return y_priv

def pure_rr_parameter(eps: float) -> float:
    return math.exp(eps) / (math.exp(eps) + 1.0)

def approx_rr_parameter(eps: float, delta: float) -> float:
    return (math.exp(eps) + delta) / (math.exp(eps) + 1.0)

def local_bounds(center: np.ndarray, radius: float) -> List[Tuple[float, float]]:
    """Generates bounding box for local optimization."""
    return [(float(c - radius), float(c + radius)) for c in center]

# ---------------------------
# True SVM parameter for the Gaussian design
# ---------------------------
def true_theta_gaussian(mu: np.ndarray, Sigma: np.ndarray) -> np.ndarray:
    Sigma_inv = np.linalg.inv(Sigma)
    m = float(np.sqrt(mu @ Sigma_inv @ mu))

    def mills_ratio(a: float) -> float:
        return float(norm.pdf(a) / norm.cdf(a))

    a_star = brentq(lambda a: mills_ratio(a) - m, -20.0, 20.0)
    w_star = Sigma_inv @ mu / (m * (a_star + m))
    return np.r_[0.0, w_star]

THETA_TRUE = true_theta_gaussian(SIM_MU, SIM_SIGMA)

def posterior_q_balanced_gaussian(X_no_intercept: np.ndarray, mu: np.ndarray, Sigma: np.ndarray) -> np.ndarray:
    Sigma_inv = np.linalg.inv(Sigma)
    logit = 2.0 * (X_no_intercept @ Sigma_inv @ mu)
    return 1.0 / (1.0 + np.exp(-logit))

def verify_symmetric_approx_candidate(
    mu: np.ndarray,
    Sigma: np.ndarray,
    theta_true: np.ndarray,
    delta: float,
    eps_grid: Sequence[float],
    mc_n: int = 250000,
    seed: int = 12345,
) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    d = len(mu)
    y = np.where(rng.random(mc_n) < 0.5, 1, -1)
    X0 = rng.multivariate_normal(mean=np.zeros(d), cov=Sigma, size=mc_n) + y[:, None] * mu
    X = np.c_[np.ones(mc_n), X0]
    q = posterior_q_balanced_gaussian(X0, mu, Sigma)
    z = X @ theta_true
    A = (1.0 - z > 0.0).astype(float)
    B = (1.0 + z > 0.0).astype(float)
    common_weight = (A + B) ** 2 * np.sum(X * X, axis=1)

    rows = []
    for eps in eps_grid:
        p_sym = approx_rr_parameter(eps, delta)
        candidates = {
            "symmetric": (p_sym, p_sym),
            "(1,delta)": (1.0, delta),
            "(delta,1)": (delta, 1.0),
        }
        for label, (pm, pp) in candidates.items():
            Delta = pm + pp - 1.0
            numer = pp * (1.0 - pp) * q + pm * (1.0 - pm) * (1.0 - q)
            T_value = float(np.mean(numer / (Delta * Delta) * common_weight))
            rows.append({"epsilon": eps, "candidate": label, "T_value": T_value})
    df = pd.DataFrame(rows)
    return df

# ---------------------------
# Simulation study
# ---------------------------
def simulation_one(seed: int) -> Dict:
    rng = np.random.default_rng(10_000 + seed)
    d = len(SIM_MU)

    y_true = np.where(rng.random(SIM_N) < 0.5, 1, -1)
    X0 = rng.multivariate_normal(mean=np.zeros(d), cov=SIM_SIGMA, size=SIM_N) + y_true[:, None] * SIM_MU
    X = np.c_[np.ones(SIM_N), X0]

    # NP and RRSVM Baseline parameters
    h_np = (d / SIM_N) ** 0.25
    lam_np = 0.0

    # Proposed ORRSVM parameters for Simulation
    h_dp = H_SCALE * (SIM_N ** (-1 / 3))
    lam_dp = LAM_SCALE * (SIM_N ** (-2 / 3))

    # Non-private fit (Benchmark ONLY)
    np_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X, y_true, args=(h_np, lam_np))
    _, Sigma_np = sandwich_covariance(np_fit.hess, np_fit.score_i)

    out = {
        "np_mse": float(np.sum((np_fit.theta - THETA_TRUE) ** 2)),
        "np_acp": ci_coverage(np_fit.theta, THETA_TRUE, Sigma_np, SIM_N),
        "pure": {},
        "approx": {},
    }

    for eps in EPS_GRID:
        # 1. Pure epsilon-LabelDP
        p_pure = pure_rr_parameter(eps)
        y_priv_pure = privatize_labels(y_true, p_pure, p_pure, rng)

        rr_pure_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X, y_priv_pure, args=(h_np, lam_np))
        center_pure = rr_pure_fit.theta.copy()
        _, Sigma_rr_pure = sandwich_covariance(rr_pure_fit.hess, rr_pure_fit.score_i)

        dp_pure_fit = fit_lbfgs(
            dpsvm_components,
            center_pure,
            X,
            y_priv_pure,
            args=(p_pure, p_pure, h_dp, lam_dp),
            bounds=local_bounds(center_pure, LOCAL_BOX_RADIUS),
        )
        _, Sigma_dp_pure = sandwich_covariance(dp_pure_fit.hess, dp_pure_fit.score_i)

        out["pure"][eps] = {
            "rr_mse": float(np.sum((rr_pure_fit.theta - THETA_TRUE) ** 2)),
            "rr_acp": ci_coverage(rr_pure_fit.theta, THETA_TRUE, Sigma_rr_pure, SIM_N),
            "dp_mse": float(np.sum((dp_pure_fit.theta - THETA_TRUE) ** 2)),
            "dp_acp": ci_coverage(dp_pure_fit.theta, THETA_TRUE, Sigma_dp_pure, SIM_N),
        }

        # 2. Approximate (epsilon, delta)-LabelDP
        sym_p = approx_rr_parameter(eps, DELTA_APPROX)
        candidates = [
            (1.0, DELTA_APPROX),
            (DELTA_APPROX, 1.0),
            (sym_p, sym_p)
        ]
        
        best_trace = float('inf')
        best_approx_res = None

        for (pm, pp) in candidates:
            y_priv_cand = privatize_labels(y_true, pm, pp, rng)
            
            # Guardrail against vanishing variance (class extinction)
            counts = np.unique(y_priv_cand, return_counts=True)[1]
            if len(counts) < 2 or min(counts) < 5:
                continue
            
            rr_cand_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X, y_priv_cand, args=(h_np, lam_np))
            center_cand = rr_cand_fit.theta.copy()
            
            dp_cand_fit = fit_lbfgs(
                dpsvm_components,
                center_cand,
                X,
                y_priv_cand,
                args=(pm, pp, h_dp, lam_dp),
                bounds=local_bounds(center_cand, LOCAL_BOX_RADIUS)
            )

            V_hat_cand, Sigma_dp_cand = sandwich_covariance(dp_cand_fit.hess, dp_cand_fit.score_i)
            trace_cand = float(np.trace(V_hat_cand))

            if trace_cand < best_trace:
                best_trace = trace_cand
                _, Sigma_rr_cand = sandwich_covariance(rr_cand_fit.hess, rr_cand_fit.score_i)
                best_approx_res = {
                    "rr_mse": float(np.sum((rr_cand_fit.theta - THETA_TRUE) ** 2)),
                    "rr_acp": ci_coverage(rr_cand_fit.theta, THETA_TRUE, Sigma_rr_cand, SIM_N),
                    "dp_mse": float(np.sum((dp_cand_fit.theta - THETA_TRUE) ** 2)),
                    "dp_acp": ci_coverage(dp_cand_fit.theta, THETA_TRUE, Sigma_dp_cand, SIM_N),
                }

        out["approx"][eps] = best_approx_res

    return out

def run_simulation() -> Tuple[pd.DataFrame, pd.DataFrame, Dict]:
    start = time.time()
    rows: List[Dict] = []
    for i in range(1, SIM_B + 1):
        rows.append(simulation_one(i - 1))
        if i % 50 == 0 or i == SIM_B:
            print(f"[simulation] finished {i}/{SIM_B} replications")

    def agg_table(regime: str) -> pd.DataFrame:
        data = []
        np_mse = float(np.mean([r["np_mse"] for r in rows]))
        np_acp = float(np.mean([r["np_acp"] for r in rows]))
        np_mse_sd = float(np.std([r["np_mse"] for r in rows], ddof=1))
        np_acp_sd = float(np.std([r["np_acp"] for r in rows], ddof=1))
        for eps in EPS_GRID:
            rr_mse = np.array([r[regime][eps]["rr_mse"] for r in rows], dtype=float)
            rr_acp = np.array([r[regime][eps]["rr_acp"] for r in rows], dtype=float)
            dp_mse = np.array([r[regime][eps]["dp_mse"] for r in rows], dtype=float)
            dp_acp = np.array([r[regime][eps]["dp_acp"] for r in rows], dtype=float)
            data.append(
                {
                    "epsilon": eps,
                    "NP_log10_MSE": math.log10(np_mse),
                    "NP_log10_MSE_sd_raw": np_mse_sd,
                    "NP_ACP": np_acp,
                    "NP_ACP_sd": np_acp_sd,
                    "RR_log10_MSE": math.log10(float(rr_mse.mean())),
                    "RR_log10_MSE_sd_raw": float(rr_mse.std(ddof=1)),
                    "RR_ACP": float(rr_acp.mean()),
                    "RR_ACP_sd": float(rr_acp.std(ddof=1)),
                    "DPSVM_log10_MSE": math.log10(float(dp_mse.mean())),
                    "DPSVM_log10_MSE_sd_raw": float(dp_mse.std(ddof=1)),
                    "DPSVM_ACP": float(dp_acp.mean()),
                    "DPSVM_ACP_sd": float(dp_acp.std(ddof=1)),
                }
            )
        return pd.DataFrame(data)

    pure_df = agg_table("pure")
    approx_df = agg_table("approx")
    meta = {
        "simulation_seconds": time.time() - start,
        "theta_true": THETA_TRUE.tolist(),
        "sim_workers": SIM_WORKERS,
        "sim_B": SIM_B,
    }
    return pure_df, approx_df, meta


# ---------------------------
# Real-data study (Spambase N=4601)
# ---------------------------
def load_spambase_data():
    try:
        from sklearn.datasets import fetch_openml
        print("Downloading/Loading Spambase dataset (N=4601)...")
        data = fetch_openml(name='spambase', version=1, as_frame=False, parser='auto')
        X_raw = data.data
        y_raw = np.where(data.target.astype(int) == 1, 1, -1)
        dataset_name = "Spambase"
    except Exception as e:
        print(f"Network error fetching Spambase: {e}")
        print("Falling back to make_classification (N=4601)...")
        from sklearn.datasets import make_classification
        X_raw, y_raw = make_classification(n_samples=4601, n_features=57, n_informative=10, random_state=2026)
        y_raw = np.where(y_raw == 1, 1, -1)
        dataset_name = "Synthetic (Spambase Fallback)"
    
    return X_raw, y_raw, dataset_name

REAL_X_RAW, REAL_Y, DATASET_NAME = load_spambase_data()

def accuracy(theta: np.ndarray, X: np.ndarray, y: np.ndarray) -> float:
    pred = np.where(X @ theta >= 0.0, 1, -1)
    return float(np.mean(pred == y))

def real_data_one(split_id: int, train_idx: np.ndarray, test_idx: np.ndarray) -> Dict:
    X_train_raw = REAL_X_RAW[train_idx]
    X_test_raw = REAL_X_RAW[test_idx]
    y_train = REAL_Y[train_idx]
    y_test = REAL_Y[test_idx]

    scaler = StandardScaler().fit(X_train_raw)
    pca = PCA(n_components=REAL_PCA_COMPONENTS, random_state=2026)
    
    X_train_scaled = pca.fit_transform(scaler.transform(X_train_raw))
    X_test_scaled = pca.transform(scaler.transform(X_test_raw))

    X_train = np.c_[np.ones(X_train_scaled.shape[0]), X_train_scaled]
    X_test = np.c_[np.ones(X_test_scaled.shape[0]), X_test_scaled]

    d = X_train.shape[1] - 1
    n_train = X_train.shape[0]

    # NP and RRSVM Baseline parameters
    h_np = (d / n_train) ** 0.25
    lam_np = 0.0

    # Proposed ORRSVM parameters for Real Data
    h_dp = REAL_H_SCALE * (n_train ** (-1 / 3))
    lam_dp = REAL_LAM_SCALE * (n_train ** (-2 / 3))

    np_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X_train, y_train, args=(h_np, lam_np))

    rng = np.random.default_rng(200_000 + split_id)
    out = {"np_acc": accuracy(np_fit.theta, X_test, y_test), "pure": {}, "approx": {}}

    for eps in EPS_GRID:
        # 1. Pure epsilon-LabelDP
        p_pure = pure_rr_parameter(eps)
        y_priv_pure = privatize_labels(y_train, p_pure, p_pure, rng)

        rr_pure_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X_train, y_priv_pure, args=(h_np, lam_np))
        center_pure = rr_pure_fit.theta.copy()
        
        dp_pure_fit = fit_lbfgs(
            dpsvm_components,
            center_pure,
            X_train,
            y_priv_pure,
            args=(p_pure, p_pure, h_dp, lam_dp),
            bounds=local_bounds(center_pure, REAL_LOCAL_BOX_RADIUS), 
        )

        out["pure"][eps] = {
            "rr_acc": accuracy(rr_pure_fit.theta, X_test, y_test),
            "dp_acc": accuracy(dp_pure_fit.theta, X_test, y_test),
        }

        # 2. Approximate (epsilon, delta)-LabelDP
        sym_p = approx_rr_parameter(eps, DELTA_APPROX)
        candidates = [
            (1.0, DELTA_APPROX),
            (DELTA_APPROX, 1.0),
            (sym_p, sym_p)
        ]
        
        best_trace = float('inf')
        best_approx_res = None

        for (pm, pp) in candidates:
            y_priv_cand = privatize_labels(y_train, pm, pp, rng)
            
            # Guardrail against vanishing variance
            counts = np.unique(y_priv_cand, return_counts=True)[1]
            if len(counts) < 2 or min(counts) < 5:
                continue
                
            rr_cand_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X_train, y_priv_cand, args=(h_np, lam_np))
            center_cand = rr_cand_fit.theta.copy()
            
            dp_cand_fit = fit_lbfgs(
                dpsvm_components,
                center_cand,
                X_train,
                y_priv_cand,
                args=(pm, pp, h_dp, lam_dp),
                bounds=local_bounds(center_cand, REAL_LOCAL_BOX_RADIUS), 
            )

            V_hat_cand, _ = sandwich_covariance(dp_cand_fit.hess, dp_cand_fit.score_i)
            trace_cand = float(np.trace(V_hat_cand))

            if trace_cand < best_trace:
                best_trace = trace_cand
                best_approx_res = {
                    "rr_acc": accuracy(rr_cand_fit.theta, X_test, y_test),
                    "dp_acc": accuracy(dp_cand_fit.theta, X_test, y_test),
                }

        out["approx"][eps] = best_approx_res

    return out

def run_real_data() -> Tuple[pd.DataFrame, pd.DataFrame, Dict]:
    start = time.time()
    splitter = StratifiedShuffleSplit(n_splits=REAL_SPLITS, test_size=0.30, random_state=2026)
    split_indices = list(splitter.split(REAL_X_RAW, REAL_Y))

    rows: List[Dict] = []
    for i, (tr, te) in enumerate(split_indices, 1):
        rows.append(real_data_one(i - 1, tr, te))
        if i % 10 == 0 or i == REAL_SPLITS:
            print(f"[real data] finished {i}/{REAL_SPLITS} train/test splits")

    def agg_table(regime: str) -> pd.DataFrame:
        np_acc = np.array([r["np_acc"] for r in rows], dtype=float)
        data = []
        for eps in EPS_GRID:
            rr_acc = np.array([r[regime][eps]["rr_acc"] for r in rows], dtype=float)
            dp_acc = np.array([r[regime][eps]["dp_acc"] for r in rows], dtype=float)
            data.append(
                {
                    "epsilon": eps,
                    "NP_accuracy": float(np_acc.mean()),
                    "NP_accuracy_sd": float(np_acc.std(ddof=1)),
                    "RR_accuracy": float(rr_acc.mean()),
                    "RR_accuracy_sd": float(rr_acc.std(ddof=1)),
                    "DPSVM_accuracy": float(dp_acc.mean()),
                    "DPSVM_accuracy_sd": float(dp_acc.std(ddof=1)),
                }
            )
        return pd.DataFrame(data)

    pure_df = agg_table("pure")
    approx_df = agg_table("approx")
    meta = {
        "real_seconds": time.time() - start,
        "real_splits": REAL_SPLITS,
        "real_workers": REAL_WORKERS,
        "pca_components": REAL_PCA_COMPONENTS,
        "dataset_name": DATASET_NAME,
        "n_samples": len(REAL_Y)
    }
    return pure_df, approx_df, meta


# ---------------------------
# Plotting
# ---------------------------
def plot_simulation_metric(
    pure_df: pd.DataFrame,
    approx_df: pd.DataFrame,
    metric_columns: Tuple[str, str, str],
    ylabel: str,
    filename: Path,
) -> None:
    np_col, rr_col, dp_col = metric_columns
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), sharex=True)
    for ax, df, title in zip(
        axes,
        [pure_df, approx_df],
        [r"$\delta=0$ (pure $\varepsilon$-LabelDP)", rf"$\delta={DELTA_APPROX}$ (($\varepsilon,\delta$)-LabelDP)"],
    ):
        ax.plot(df["epsilon"], df[np_col], marker="o", label="NP")
        ax.plot(df["epsilon"], df[rr_col], marker="s", label="RR-SVM")
        ax.plot(df["epsilon"], df[dp_col], marker="^", label="DPSVM")
        ax.set_title(title)
        ax.set_xlabel(r"$\varepsilon$")
        ax.set_ylabel(ylabel)
        ax.grid(True, alpha=0.3)
    axes[0].legend(frameon=False)
    fig.tight_layout()
    fig.savefig(filename, dpi=220, bbox_inches="tight")
    plt.close(fig)

def plot_real_accuracy(
    pure_df: pd.DataFrame,
    approx_df: pd.DataFrame,
    filename: Path,
) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), sharex=True, sharey=True)
    for ax, df, title in zip(
        axes,
        [pure_df, approx_df],
        [r"$\delta=0$ (pure $\varepsilon$-LabelDP)", rf"$\delta={DELTA_APPROX}$ (($\varepsilon,\delta$)-LabelDP)"],
    ):
        ax.plot(df["epsilon"], df["NP_accuracy"], marker="o", label="NP")
        ax.plot(df["epsilon"], df["RR_accuracy"], marker="s", label="RR-SVM")
        ax.plot(df["epsilon"], df["DPSVM_accuracy"], marker="^", label="DPSVM")
        ax.set_title(title)
        ax.set_xlabel(r"$\varepsilon$")
        ax.set_ylabel("Accuracy")
        ax.grid(True, alpha=0.3)
    axes[0].legend(frameon=False)
    fig.tight_layout()
    fig.savefig(filename, dpi=220, bbox_inches="tight")
    plt.close(fig)


# ---------------------------
# Spreadsheet / table export
# ---------------------------
def write_df_to_sheet(ws, df: pd.DataFrame, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    header_row = 3
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=float(value) if isinstance(value, (np.floating, float)) else value)
    ws.freeze_panes = "A4"
    border = Border(bottom=Side(style="thin", color="BFBFBF"))
    for col in range(1, len(df.columns) + 1):
        ws.cell(row=header_row, column=col).border = border
        width = max(len(str(df.columns[col - 1])) + 2, 14)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(df), min_col=1, max_col=len(df.columns)):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if "accuracy" in str(ws.cell(header_row, cell.column).value).lower() or "acp" in str(ws.cell(header_row, cell.column).value).lower():
                    cell.number_format = "0.000"
                elif "epsilon" in str(ws.cell(header_row, cell.column).value).lower():
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "0.000"

def build_workbook(
    sim_pure_df: pd.DataFrame,
    sim_approx_df: pd.DataFrame,
    real_pure_df: pd.DataFrame,
    real_approx_df: pd.DataFrame,
    candidate_df: pd.DataFrame,
    settings: Dict,
    filepath: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulation_pure"
    write_df_to_sheet(ws, sim_pure_df, "Simulation results: pure epsilon-LabelDP")

    ws2 = wb.create_sheet("Simulation_approx")
    write_df_to_sheet(ws2, sim_approx_df, f"Simulation results: (epsilon, delta)-LabelDP, delta={DELTA_APPROX}")

    ws3 = wb.create_sheet("RealData_pure")
    write_df_to_sheet(ws3, real_pure_df, "Real-data accuracy: pure epsilon-LabelDP")

    ws4 = wb.create_sheet("RealData_approx")
    write_df_to_sheet(ws4, real_approx_df, f"Real-data accuracy: (epsilon, delta)-LabelDP, delta={DELTA_APPROX}")

    ws5 = wb.create_sheet("Approx_T_check")
    write_df_to_sheet(ws5, candidate_df, "Population T-check on the chosen Gaussian scenario")

    ws6 = wb.create_sheet("Settings")
    ws6["A1"] = "Setting"
    ws6["B1"] = "Value"
    ws6["A1"].font = ws6["B1"].font = Font(bold=True, color="FFFFFF")
    ws6["A1"].fill = ws6["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    settings_rows = [
        ("Simulation n", SIM_N),
        ("Simulation B", SIM_B),
        ("Epsilon grid", ", ".join(f"{e:.2f}" if e < 1 else f"{e:.1f}" for e in EPS_GRID)),
        ("Delta (approx)", DELTA_APPROX),
        ("Simulation H_SCALE", H_SCALE),
        ("Simulation LAM_SCALE", LAM_SCALE),
        ("Simulation Box Radius", LOCAL_BOX_RADIUS),
        ("Real Data H_SCALE", REAL_H_SCALE),
        ("Real Data LAM_SCALE", REAL_LAM_SCALE),
        ("Real Data Box Radius", REAL_LOCAL_BOX_RADIUS),
        ("Simulation mu", json.dumps(SIM_MU.tolist())),
        ("Simulation Sigma", json.dumps(SIM_SIGMA.tolist())),
        ("True theta", json.dumps(THETA_TRUE.tolist())),
        ("Real dataset", settings.get("dataset_name", DATASET_NAME)),
        ("Real Data Samples (N)", settings.get("n_samples", len(REAL_Y))),
        ("Real PCA components", settings.get("pca_components", REAL_PCA_COMPONENTS)),
        ("Simulation runtime (s)", round(settings["simulation_seconds"], 2)),
        ("Real-data runtime (s)", round(settings["real_seconds"], 2)),
        ("Total runtime (s)", round(settings["total_seconds"], 2)),
    ]
    for i, (k, v) in enumerate(settings_rows, start=2):
        ws6[f"A{i}"] = k
        ws6[f"B{i}"] = v
    ws6.column_dimensions["A"].width = 32
    ws6.column_dimensions["B"].width = 80

    wb.save(filepath)


# ---------------------------
# Main driver
# ---------------------------
def main() -> None:
    total_start = time.time()

    print("Running T-optimality check for the approximate mechanism...")
    candidate_df = verify_symmetric_approx_candidate(SIM_MU, SIM_SIGMA, THETA_TRUE, DELTA_APPROX, EPS_GRID)
    candidate_df.to_csv(OUTPUT_DIR / "approx_candidate_T_check.csv", index=False)

    print("Running simulation study...")
    sim_pure_df, sim_approx_df, sim_meta = run_simulation()
    sim_pure_df.to_csv(OUTPUT_DIR / "simulation_pure_table.csv", index=False)
    sim_approx_df.to_csv(OUTPUT_DIR / "simulation_approx_table.csv", index=False)

    print(f"Running real-data study on {DATASET_NAME} (N={len(REAL_Y)})...")
    real_pure_df, real_approx_df, real_meta = run_real_data()
    real_pure_df.to_csv(OUTPUT_DIR / "realdata_pure_table.csv", index=False)
    real_approx_df.to_csv(OUTPUT_DIR / "realdata_approx_table.csv", index=False)

    print("Creating figures...")
    plot_simulation_metric(
        sim_pure_df,
        sim_approx_df,
        metric_columns=("NP_log10_MSE", "RR_log10_MSE", "DPSVM_log10_MSE"),
        ylabel=r"$\log_{10}(\mathrm{MSE})$",
        filename=OUTPUT_DIR / "simulation_log10_mse.png",
    )
    plot_simulation_metric(
        sim_pure_df,
        sim_approx_df,
        metric_columns=("NP_ACP", "RR_ACP", "DPSVM_ACP"),
        ylabel="ACP",
        filename=OUTPUT_DIR / "simulation_acp.png",
    )
    plot_real_accuracy(real_pure_df, real_approx_df, OUTPUT_DIR / "realdata_accuracy.png")

    total_seconds = time.time() - total_start
    settings = {
        **sim_meta,
        **real_meta,
        "total_seconds": total_seconds,
    }
    with open(OUTPUT_DIR / "experiment_settings.json", "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2, ensure_ascii=False)

    summary_lines = [
        "DPSVM experiment summary",
        f"Simulation scenario: balanced Gaussian class-conditionals with mu={SIM_MU.tolist()} and Sigma={SIM_SIGMA.tolist()}",
        f"True theta: {THETA_TRUE.tolist()}",
        f"Approximate mechanism delta: {DELTA_APPROX}",
        f"Sim ORRSVM: h_n = {H_SCALE} * n^(-1/3), lambda_n = {LAM_SCALE} * n^(-2/3), Box = {LOCAL_BOX_RADIUS}",
        f"Real ORRSVM: h_n = {REAL_H_SCALE} * n^(-1/3), lambda_n = {REAL_LAM_SCALE} * n^(-2/3), Box = {REAL_LOCAL_BOX_RADIUS}",
        f"Baseline (NP/RRSVM): h_np = (d/n)^(1/4), lambda_np = 0",
        f"Real-data feature strategy: PCA to {REAL_PCA_COMPONENTS} orthogonal components on {DATASET_NAME}",
        f"Simulation runtime (s): {sim_meta['simulation_seconds']:.2f}",
        f"Real-data runtime (s): {real_meta['real_seconds']:.2f}",
        f"Total runtime (s): {total_seconds:.2f}",
    ]
    with open(OUTPUT_DIR / "README_results.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(summary_lines) + "\n")

    print("Creating workbook...")
    build_workbook(
        sim_pure_df,
        sim_approx_df,
        real_pure_df,
        real_approx_df,
        candidate_df,
        settings,
        OUTPUT_DIR / "dpsvm_results_tables.xlsx",
    )

    zip_path = OUTPUT_DIR.parent / "dpsvm_experiments_artifacts.zip"
    os.system(f"cd {OUTPUT_DIR.parent} && rm -f {zip_path.name} && zip -rq {zip_path.name} {OUTPUT_DIR.name}")

    print("Done.")
    print(f"Outputs saved to: {OUTPUT_DIR}")
    print(f"Zip archive saved to: {zip_path}")


if __name__ == "__main__":
    main()