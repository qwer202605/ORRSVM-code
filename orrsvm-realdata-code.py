#!/usr/bin/env python3
"""
DPSVM Real-Data Experiment (Local CSV Version)

应用背景: 医疗临床诊断 (心血管疾病预测 Cardiovascular Disease)
隐私场景: 标签差分隐私 (LabelDP) - 保护高度敏感的患者疾病状态
"""

import os
# 防止多进程环境下的 BLAS 线程过载
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("VECLIB_MAXIMUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")

from dataclasses import dataclass
import json
import math
import time
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from scipy.optimize import minimize
from scipy.stats import norm

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sklearn.model_selection import StratifiedShuffleSplit
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA

# ==========================================
# 1. 请在这里修改为你本地数据集的绝对路径
# ==========================================
LOCAL_DATASET_PATH = r"C:\Users\86182\Desktop\cardio_train.csv"

# ==========================================
# 2. 全局与实验参数配置
# ==========================================
EPS_GRID: List[float] = [0.05, 0.06, 0.07, 0.08, 0.09, 0.10, 0.50, 0.70, 1.00]
DELTA_APPROX: float = 1e-4

REAL_SPLITS: int = 10  # 评估划分次数 (7万样本较大，设为10次以节省时间，可依算力调大)
REAL_LOCAL_BOX_RADIUS: float = 4.0
REAL_H_SCALE: float = 0.80
REAL_LAM_SCALE: float = 0.10
REAL_PCA_COMPONENTS: int = 5  # PCA 降维保留的最重要主成分数量

OUTPUT_DIR = Path("C:\\Users\\86182\\Desktop\\新建文件夹 (3)")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SQRT_2PI = math.sqrt(2.0 * math.pi)

# ---------------------------
# 核心数值与分布计算
# ---------------------------
def phi(z: np.ndarray) -> np.ndarray:
    return np.exp(-0.5 * z * z) / SQRT_2PI

def Phi(z: np.ndarray) -> np.ndarray:
    return norm.cdf(z)

def smoothed_hinge(u: np.ndarray, h: float) -> np.ndarray:
    z = u / h
    return u * Phi(z) + h * phi(z)

@dataclass
class FitResult:
    theta: np.ndarray
    loss: float
    hess: np.ndarray
    score_i: np.ndarray

# ---------------------------
# 模型目标函数与导数 (修正的平滑 SVM)
# ---------------------------
def svm_components(theta: np.ndarray, X: np.ndarray, y: np.ndarray, h: float, lam: float) -> Tuple[float, np.ndarray, np.ndarray, np.ndarray]:
    z = X @ theta
    u = 1.0 - y * z
    cdf = Phi(u / h)
    pdf = phi(u / h)

    loss = float(smoothed_hinge(u, h).mean() + 0.5 * lam * np.dot(theta[1:], theta[1:]))
    grad = -(X * (y * cdf)[:, None]).mean(axis=0)
    grad[1:] += lam * theta[1:]

    hess = (X.T * (pdf / h)).dot(X) / X.shape[0]
    hess[1:, 1:] += lam * np.eye(X.shape[1] - 1)

    score_i = -(X * (y * cdf)[:, None])
    return loss, grad, hess, score_i

def dpsvm_components(theta: np.ndarray, X: np.ndarray, y_priv: np.ndarray, p_minus: float, p_plus: float, h: float, lam: float) -> Tuple[float, np.ndarray, np.ndarray, np.ndarray]:
    Delta = p_minus + p_plus - 1.0
    z = X @ theta
    u_minus = 1.0 - z
    u_plus = 1.0 + z

    cdf_minus = Phi(u_minus / h)
    cdf_plus = Phi(u_plus / h)
    pdf_minus = phi(u_minus / h)
    pdf_plus = phi(u_plus / h)

    loss_minus = smoothed_hinge(u_minus, h)
    loss_plus = smoothed_hinge(u_plus, h)

    pos = y_priv == 1
    neg = ~pos

    loss_i = np.empty_like(z)
    loss_i[pos] = (p_minus * loss_minus[pos] - (1.0 - p_plus) * loss_plus[pos]) / Delta
    loss_i[neg] = (p_plus * loss_plus[neg] - (1.0 - p_minus) * loss_minus[neg]) / Delta
    loss = float(loss_i.mean() + 0.5 * lam * np.dot(theta[1:], theta[1:]))

    weight = np.empty_like(z)
    weight[pos] = -(p_minus * cdf_minus[pos] + (1.0 - p_plus) * cdf_plus[pos]) / Delta
    weight[neg] = (p_plus * cdf_plus[neg] + (1.0 - p_minus) * cdf_minus[neg]) / Delta
    grad = (X * weight[:, None]).mean(axis=0)
    grad[1:] += lam * theta[1:]

    curv = np.empty_like(z)
    curv[pos] = (p_minus * pdf_minus[pos] - (1.0 - p_plus) * pdf_plus[pos]) / (h * Delta)
    curv[neg] = (p_plus * pdf_plus[neg] - (1.0 - p_minus) * pdf_minus[neg]) / (h * Delta)
    hess = (X.T * curv).dot(X) / X.shape[0]
    hess[1:, 1:] += lam * np.eye(X.shape[1] - 1)

    score_i = X * weight[:, None]
    return loss, grad, hess, score_i

def fit_lbfgs(
    components_func,
    theta0: np.ndarray,
    X: np.ndarray,
    y: np.ndarray,
    args: Tuple,
    bounds=None,
    maxiter: int = 300,
) -> FitResult:
    objective = lambda th: components_func(th, X, y, *args)[0]
    gradient = lambda th: components_func(th, X, y, *args)[1]
    res = minimize(
        objective,
        theta0,
        jac=gradient,
        method="L-BFGS-B",
        bounds=bounds,
        options={"maxiter": maxiter, "ftol": 1e-12, "gtol": 1e-8},
    )
    loss, _, hess, score_i = components_func(res.x, X, y, *args)
    return FitResult(theta=res.x, loss=float(loss), hess=hess, score_i=score_i)

def sandwich_covariance(hess: np.ndarray, score_i: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    n = score_i.shape[0]
    centered = score_i - score_i.mean(axis=0, keepdims=True)
    V_hat = centered.T @ centered / n
    H = 0.5 * (hess + hess.T)
    H_inv = np.linalg.pinv(H)
    Sigma_hat = H_inv @ V_hat @ H_inv
    return V_hat, Sigma_hat

def privatize_labels(y_true: np.ndarray, p_minus: float, p_plus: float, rng: np.random.Generator) -> np.ndarray:
    y_priv = y_true.copy()
    neg = y_true == -1
    pos = ~neg
    y_priv[neg] = np.where(rng.random(neg.sum()) < p_minus, -1, 1)
    y_priv[pos] = np.where(rng.random(pos.sum()) < p_plus, 1, -1)
    return y_priv

def pure_rr_parameter(eps: float) -> float:
    return math.exp(eps) / (math.exp(eps) + 1.0)

def approx_rr_parameter(eps: float, delta: float) -> float:
    return (math.exp(eps) + delta) / (math.exp(eps) + 1.0)

def local_bounds(center: np.ndarray, radius: float) -> List[Tuple[float, float]]:
    return [(float(c - radius), float(c + radius)) for c in center]

def accuracy(theta: np.ndarray, X: np.ndarray, y: np.ndarray) -> float:
    pred = np.where(X @ theta >= 0.0, 1, -1)
    return float(np.mean(pred == y))

# ---------------------------
# 加载本地数据集
# ---------------------------
def load_local_dataset(filepath: str):
    print(f"正在从本地读取数据集: {filepath}")
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"找不到文件: {filepath}。请检查 LOCAL_DATASET_PATH 是否正确。")
    
    # Kaggle 的 cardio 数据集通常以 ';' 分隔。如果报错，可以尝试改为 sep=','
    df = pd.read_csv(filepath, sep=';')
    
    # 尝试处理以防它是逗号分隔的
    if len(df.columns) == 1:
        df = pd.read_csv(filepath, sep=',')
        
    print(f"成功读取数据！总行数: {len(df)}, 总列数: {len(df.columns)}")
    
    # 目标列通常叫 'cardio'
    target_col = 'cardio'
    if target_col not in df.columns:
        raise ValueError(f"数据集中没有找到名为 '{target_col}' 的目标列，请检查数据。可用的列有: {df.columns.tolist()}")

    # 将 0 转换为 -1，1 保持为 1
    y_raw = df[target_col].values
    y_raw = np.where(y_raw == 1, 1, -1)
    
    # 移除标签列和无用的 'id' 列
    drop_cols = [target_col]
    if 'id' in df.columns:
        drop_cols.append('id')
    
    X_raw = df.drop(columns=drop_cols).values
    dataset_name = "Local_Cardiovascular_Disease"
    
    print(f"特征提取完毕。特征数量: {X_raw.shape[1]}")
    print(f"正负类分布: {np.mean(y_raw==1):.2%} 正类 (Cardio=1) vs {np.mean(y_raw==-1):.2%} 负类 (Cardio=0)")
    
    return X_raw, y_raw, dataset_name

# 全局变量占位，稍后在 main 中初始化
REAL_X_RAW = None
REAL_Y = None
DATASET_NAME = None

# ---------------------------
# 真实数据实验核心流程
# ---------------------------
def real_data_one(split_id: int, train_idx: np.ndarray, test_idx: np.ndarray) -> Dict:
    X_train_raw = REAL_X_RAW[train_idx]
    X_test_raw = REAL_X_RAW[test_idx]
    y_train = REAL_Y[train_idx]
    y_test = REAL_Y[test_idx]

    # 标准化
    scaler = StandardScaler().fit(X_train_raw)
    X_train_scaled = scaler.transform(X_train_raw)
    X_test_scaled = scaler.transform(X_test_raw)

    # PCA 降维，提取最重要的特征
    pca = PCA(n_components=REAL_PCA_COMPONENTS, random_state=2026)
    X_train_scaled = pca.fit_transform(X_train_scaled)
    X_test_scaled = pca.transform(X_test_scaled)

    # 增加截距项
    X_train = np.c_[np.ones(X_train_scaled.shape[0]), X_train_scaled]
    X_test = np.c_[np.ones(X_test_scaled.shape[0]), X_test_scaled]

    d = X_train.shape[1] - 1
    n_train = X_train.shape[0]

    h_np = (d / n_train) ** 0.25
    lam_np = 0.0

    h_dp = REAL_H_SCALE * (n_train ** (-1 / 3))
    lam_dp = REAL_LAM_SCALE * (n_train ** (-2 / 3))

    # Benchmark: 非隐私基准
    np_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X_train, y_train, args=(h_np, lam_np))

    rng = np.random.default_rng(200_000 + split_id)
    out = {"np_acc": accuracy(np_fit.theta, X_test, y_test), "pure": {}, "approx": {}}

    for eps in EPS_GRID:
        # ---- 1. ε-LabelDP (纯差分隐私) ----
        p_pure = pure_rr_parameter(eps)
        y_priv_pure = privatize_labels(y_train, p_pure, p_pure, rng)

        rr_pure_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X_train, y_priv_pure, args=(h_np, lam_np))
        center_pure = rr_pure_fit.theta.copy()
        
        dp_pure_fit = fit_lbfgs(
            dpsvm_components,
            center_pure,
            X_train,
            y_priv_pure,
            args=(p_pure, p_pure, h_dp, lam_dp),
            bounds=local_bounds(center_pure, REAL_LOCAL_BOX_RADIUS), 
        )

        out["pure"][eps] = {
            "rr_acc": accuracy(rr_pure_fit.theta, X_test, y_test),
            "dp_acc": accuracy(dp_pure_fit.theta, X_test, y_test),
        }

        # ---- 2. (ε, δ)-LabelDP (近似差分隐私) ----
        sym_p = approx_rr_parameter(eps, DELTA_APPROX)
        candidates = [
            (1.0, DELTA_APPROX),
            (DELTA_APPROX, 1.0),
            (sym_p, sym_p)
        ]
        
        best_trace = float('inf')
        best_approx_res = None

        for (pm, pp) in candidates:
            y_priv_cand = privatize_labels(y_train, pm, pp, rng)
            
            counts = np.unique(y_priv_cand, return_counts=True)[1]
            if len(counts) < 2 or min(counts) < 5:
                continue
                
            rr_cand_fit = fit_lbfgs(svm_components, np.zeros(d + 1), X_train, y_priv_cand, args=(h_np, lam_np))
            center_cand = rr_cand_fit.theta.copy()
            
            dp_cand_fit = fit_lbfgs(
                dpsvm_components,
                center_cand,
                X_train,
                y_priv_cand,
                args=(pm, pp, h_dp, lam_dp),
                bounds=local_bounds(center_cand, REAL_LOCAL_BOX_RADIUS), 
            )

            V_hat_cand, _ = sandwich_covariance(dp_cand_fit.hess, dp_cand_fit.score_i)
            trace_cand = float(np.trace(V_hat_cand))

            if trace_cand < best_trace:
                best_trace = trace_cand
                best_approx_res = {
                    "rr_acc": accuracy(rr_cand_fit.theta, X_test, y_test),
                    "dp_acc": accuracy(dp_cand_fit.theta, X_test, y_test),
                }

        out["approx"][eps] = best_approx_res

    return out

def run_real_data() -> Tuple[pd.DataFrame, pd.DataFrame, Dict]:
    start = time.time()
    # 70% 训练集, 30% 测试集
    splitter = StratifiedShuffleSplit(n_splits=REAL_SPLITS, test_size=0.30, random_state=2026)
    split_indices = list(splitter.split(REAL_X_RAW, REAL_Y))

    rows: List[Dict] = []
    for i, (tr, te) in enumerate(split_indices, 1):
        rows.append(real_data_one(i - 1, tr, te))
        print(f"[数据实验] 已完成 {i}/{REAL_SPLITS} 次分层采样划分")

    def agg_table(regime: str) -> pd.DataFrame:
        np_acc = np.array([r["np_acc"] for r in rows], dtype=float)
        data = []
        for eps in EPS_GRID:
            rr_acc = np.array([r[regime][eps]["rr_acc"] for r in rows], dtype=float)
            dp_acc = np.array([r[regime][eps]["dp_acc"] for r in rows], dtype=float)
            data.append(
                {
                    "epsilon": eps,
                    "NP_accuracy": float(np_acc.mean()),
                    "NP_accuracy_sd": float(np_acc.std(ddof=1)),
                    "RR_accuracy": float(rr_acc.mean()),
                    "RR_accuracy_sd": float(rr_acc.std(ddof=1)),
                    "DPSVM_accuracy": float(dp_acc.mean()),
                    "DPSVM_accuracy_sd": float(dp_acc.std(ddof=1)),
                }
            )
        return pd.DataFrame(data)

    pure_df = agg_table("pure")
    approx_df = agg_table("approx")
    meta = {
        "real_seconds": time.time() - start,
        "real_splits": REAL_SPLITS,
        "dataset_name": DATASET_NAME,
        "n_samples": len(REAL_Y)
    }
    return pure_df, approx_df, meta

# ---------------------------
# 可视化与报表输出
# ---------------------------
def plot_real_accuracy(pure_df: pd.DataFrame, approx_df: pd.DataFrame, filename: Path) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(10, 4), sharex=True, sharey=True)
    for ax, df, title in zip(
        axes,
        [pure_df, approx_df],
        [r"$\delta=0$ (pure $\varepsilon$-LabelDP)", rf"$\delta={DELTA_APPROX}$ (($\varepsilon,\delta$)-LabelDP)"],
    ):
        ax.plot(df["epsilon"], df["NP_accuracy"], marker="o", label="NP")
        ax.plot(df["epsilon"], df["RR_accuracy"], marker="s", label="RR-SVM")
        ax.plot(df["epsilon"], df["DPSVM_accuracy"], marker="^", label="ORRSVM (Ours)")
        ax.set_title(title)
        ax.set_xlabel(r"$\varepsilon$")
        ax.set_ylabel("Accuracy on Test Set")
        ax.grid(True, alpha=0.3)
    axes[0].legend(frameon=False)
    fig.tight_layout()
    fig.savefig(filename, dpi=220, bbox_inches="tight")
    plt.close(fig)

def write_df_to_sheet(ws, df: pd.DataFrame, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    header_row = 3
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=float(value) if isinstance(value, (np.floating, float)) else value)
    ws.freeze_panes = "A4"
    border = Border(bottom=Side(style="thin", color="BFBFBF"))
    for col in range(1, len(df.columns) + 1):
        ws.cell(row=header_row, column=col).border = border
        width = max(len(str(df.columns[col - 1])) + 2, 14)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(df), min_col=1, max_col=len(df.columns)):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if "epsilon" in str(ws.cell(header_row, cell.column).value).lower():
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "0.000"

def build_workbook(real_pure_df: pd.DataFrame, real_approx_df: pd.DataFrame, settings: Dict, filepath: Path) -> None:
    wb = Workbook()
    ws3 = wb.active
    ws3.title = "RealData_pure"
    write_df_to_sheet(ws3, real_pure_df, "本地心血管数据 准确率: pure epsilon-LabelDP")

    ws4 = wb.create_sheet("RealData_approx")
    write_df_to_sheet(ws4, real_approx_df, f"本地心血管数据 准确率: (epsilon, delta)-LabelDP, delta={DELTA_APPROX}")

    ws6 = wb.create_sheet("Experiment_Settings")
    ws6["A1"] = "参数项"
    ws6["B1"] = "设定值"
    ws6["A1"].font = ws6["B1"].font = Font(bold=True, color="FFFFFF")
    ws6["A1"].fill = ws6["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    settings_rows = [
        ("隐私预算 Epsilon grid", ", ".join(f"{e:.2f}" for e in EPS_GRID)),
        ("松弛因子 Delta", DELTA_APPROX),
        ("评估划分次数", REAL_SPLITS),
        ("带宽系数 H_SCALE", REAL_H_SCALE),
        ("正则化系数 LAM_SCALE", REAL_LAM_SCALE),
        ("局部优化边界 Box Radius", REAL_LOCAL_BOX_RADIUS),
        ("原始数据集路径", LOCAL_DATASET_PATH),
        ("样本总数 N", settings.get("n_samples", len(REAL_Y))),
        ("PCA 保留主成分维度", REAL_PCA_COMPONENTS),
        ("运行耗时 (秒)", round(settings["real_seconds"], 2)),
    ]
    for i, (k, v) in enumerate(settings_rows, start=2):
        ws6[f"A{i}"] = k
        ws6[f"B{i}"] = v
    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["B"].width = 80

    wb.save(filepath)

# ---------------------------
# 执行入口
# ---------------------------
def main() -> None:
    global REAL_X_RAW, REAL_Y, DATASET_NAME
    
    # 1. 尝试加载本地数据
    try:
        REAL_X_RAW, REAL_Y, DATASET_NAME = load_local_dataset(LOCAL_DATASET_PATH)
    except Exception as e:
        print(f"\n[错误] 数据加载失败: {e}")
        print("请确保已修改代码第 42 行的 LOCAL_DATASET_PATH，并确认文件存在。")
        return

    print(f"\n=== 开始本地 Cardiovascular Disease 真实数据 LabelDP 实验 ===")
    print(f"数据量过大(7万)，为兼顾运行时间，当前设置采样折数 REAL_SPLITS = {REAL_SPLITS}")
    print(f"PCA 降维保留最重要的 {REAL_PCA_COMPONENTS} 个特征\n")
    
    # 2. 跑实验
    real_pure_df, real_approx_df, real_meta = run_real_data()
    
    # 3. 结果保存
    real_pure_df.to_csv(OUTPUT_DIR / "cardio_pure_table.csv", index=False)
    real_approx_df.to_csv(OUTPUT_DIR / "cardio_approx_table.csv", index=False)

    print("\n正在生成可视化图表...")
    plot_real_accuracy(real_pure_df, real_approx_df, OUTPUT_DIR / "cardio_realdata_accuracy.png")

    with open(OUTPUT_DIR / "cardio_experiment_settings.json", "w", encoding="utf-8") as f:
        json.dump(real_meta, f, indent=2, ensure_ascii=False)

    print("正在导出 Excel 汇总结果...")
    build_workbook(
        real_pure_df,
        real_approx_df,
        real_meta,
        OUTPUT_DIR / "cardio_results_summary.xlsx",
    )

    print(f"\n=== 实验完成 ===")
    print(f"所有结果与图表已保存至文件夹: {OUTPUT_DIR.resolve()}")

if __name__ == "__main__":
    main()